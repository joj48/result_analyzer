#----------------Excel---------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font
import os
#----------------PDF-------------------------------------------------------
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.platypus import TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4



def generate_excel_report(department, df, analysis_df, total_students):

    wb = Workbook()
    ws = wb.active
    ws.title = "Result Report"

    # ---------------- HEADER ----------------

    ws["A1"] = "PERFORMANCE ANALYSIS REPORT"
    ws["A1"].font = Font(size=16, bold=True)

    ws["A3"] = "Department:"
    ws["B3"] = department

    ws["A4"] = "Total Students:"
    ws["B4"] = total_students

    ws["A5"] = "Total Subjects:"
    ws["B5"] = len([col for col in df.columns if col not in ["Register No", "CGPA"]])

    # ---------------- ANALYSIS TABLE ----------------

    start_row = 7

    headers = list(analysis_df.columns)

    # Write header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Write data rows
    for row_idx, row in enumerate(analysis_df.values, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ---------------- COMPLETE STUDENT RESULT ----------------

    start_row = start_row + len(analysis_df) + 4

    ws.cell(row=start_row, column=1, value="Complete Student Result").font = Font(bold=True)

    start_row += 2

    # Header row
    for col_num, column in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_num, value=column).font = Font(bold=True)

    # Student data
    for row_idx, row in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ---------------- SAVE FILE ----------------

    os.makedirs("outputs", exist_ok=True)
    output_path = os.path.join("outputs", f"{department}_Report.xlsx")

    wb.save(output_path)

    return output_path






def generate_pdf_report(department, df, analysis_df, total_students):

    os.makedirs("outputs", exist_ok=True)
    output_path = os.path.join("outputs", f"{department}_Report.pdf")

    doc = SimpleDocTemplate(output_path, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()

    # ---------------- HEADER ----------------

    elements.append(Paragraph("PERFORMANCE ANALYSIS REPORT", styles["Heading1"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Department: {department}", styles["Normal"]))
    elements.append(Paragraph(f"Total Students: {total_students}", styles["Normal"]))
    elements.append(
        Paragraph(
            f"Total Subjects: {len([col for col in df.columns if col not in ['Register No', 'CGPA']])}",
            styles["Normal"]
        )
    )
    elements.append(Spacer(1, 20))

    # ---------------- TOPPERS ----------------

    elements.append(Paragraph("Top 3 Students:", styles["Heading2"]))
    elements.append(Spacer(1, 12))

    # Get toppers directly from df
    top_students = df.sort_values(by="CGPA", ascending=False).head(3)

    for _, row in top_students.iterrows():
        elements.append(
            Paragraph(f"{row['Register No']} — CGPA: {row['CGPA']}", styles["Normal"])
        )

    elements.append(Spacer(1, 20))

    # ---------------- ANALYSIS TABLE ----------------

    elements.append(Paragraph("Subject-wise Performance", styles["Heading2"]))
    elements.append(Spacer(1, 12))

    table_data = [analysis_df.columns.tolist()] + analysis_df.values.tolist()

    table = Table(table_data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)

    from reportlab.platypus import PageBreak

    elements.append(PageBreak())

    elements.append(Paragraph("Complete Student Result", styles["Heading2"]))
    elements.append(Spacer(1, 12))

    # Prepare full student table
    student_table_data = [df.columns.tolist()] + df.values.tolist()

    student_table = Table(student_table_data, repeatRows=1)

    student_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
    ]))

    elements.append(student_table)

    doc.build(elements)

    return output_path
